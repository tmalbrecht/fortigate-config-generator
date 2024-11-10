from datetime import datetime
import os
from jinja2 import Environment, FileSystemLoader
import openpyxl


# Create python dictionary from xlsx file
def get_config_items_xlsx(variables):
    try:
        workbook = openpyxl.load_workbook(variables)

        config_items = {}

        # add timestamp
        config_items["time_now"] = get_time()

        # add all addresses config
        sheet = workbook["addresses"]
        config_items["addresses"] = []
        for row in sheet.iter_rows(
            min_row=1, max_row=sheet.max_row, min_col=1, max_col=2, values_only=True
        ):
            if row[0] != "name":
                config_items["addresses"].append({"name": row[0], "value": row[1]})

        return config_items

    except FileNotFoundError:
        print(f"The file {variables} does not exist.")
    except PermissionError:
        print(
            f"\nPermission denied when trying to open {variables}. \nPlease close the xlsx file so the script can access it."
        )
    except Exception as e:
        print(f"An error occurred when trying to open {variables}: {e}")


# Get the current local date/time and format the object to a string in a readable format
def get_time():
    time = datetime.now()
    time = time.strftime("%Y-%m-%d_%Hh%Mm")
    return time


if __name__ == "__main__":
    # Check if needed directories exist, if not create them
    os.makedirs("variables/", exist_ok=True)
    os.makedirs("generated_config/", exist_ok=True)
    os.makedirs("templates/", exist_ok=True)

    # Name xlsx file
    variables = "variables/variables.xlsx"

    # Create python dictionary from xlsx file and given secrets
    config_itmes = get_config_items_xlsx(variables)

    # Create jinja environment and load the template file
    env = Environment(
        loader=FileSystemLoader("."), trim_blocks=True, lstrip_blocks=True
    )
    template = env.get_template("templates/address_subnet.j2")

    # Generate config by rendering the dictionary that was created from the xlsx file, after create txt file
    ### put template in name  ###
    if config_itmes != None:
        file_name = f"generated_config/config_{get_time()}.conf"
        config = template.render(config_itmes)
        with open(file_name, "w") as w:
            w.write(config)
            print()
            print("*" * 60)
            print(
                f"Config file has been generated."
            )
            print("*" * 60)
